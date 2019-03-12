# Work with Python 3.6
import discord
import discord.user
import praw #reddit API Found Here - https://praw.readthedocs.io/en/latest/index.html
import sports  #Sports API Found Here - https://pypi.org/project/sports.py/
import openpyxl #API Found Here - https://openpyxl.readthedocs.io/en/stable/
import datetime
import random
import time
import asyncio
import emoji
import ast

TOKEN = 'token'

# Spreadsheet path
KnackchatSpreadsheetPath = r'...'

#Gets filled in later
PersonClassDict = {}

gameList = ["That Bloomin' Feeling", "Fightcloud", "Sweet Victory", "Nik-0", "Xenogrove", "Smash & Grabbers", "Quilko's Song", 
            "Of A Feather", "Rubbernecking", "Oh Snap!", "Fingers Crossed", "WWE Mayhem", "The Wall VR", 
            "Injustice 2 Mobile", "MK11", "Vindicator"]

owHeros = ['Ana', 'Bastion', 'Brigitte', 'D.Va', 'Doomfist', 'Genji', 'Hanzo', 'Junkrat', 'Lucio', 'McCree', 'Mei', 'Mercy', 'Moria', 
           'Orisa', 'Pharah', 'Reaper', 'Reinhardt', 'Roadhog', 'Solder: 76', 'Sombra', 'Symmetra', 'Torbjorn', 'Tracer', 'Widowmaker',
           'Winston', 'Zarya', 'Zenyatta']

reddit = praw.Reddit(client_id='clientid', client_secret='clientsecret', username='redditusername', password='redditpassword', user_agent='useragent')
cuteSubreddits = ['Rabbits', 'aww', 'corgi', 'BabyCorgis', 'Incorgnito', 'cats', 'rarepuppers', 'Pigtures', 'Catloaf', 'tuckedinkitties', 'babygoats', 'TuckedInPuppies',
                  'SupermodelCats', 'Superbowl', 'StuffOnCats', 'redpandas', 'duck', 'puppies', 'kittens', 'hamsters', 'Floof', 'Panda_Gifs',
                  'babyduckgifs', 'aww_gifs', 'babyelephantgifs']

commandDict = {"!roll [int X, int Y = 0]" : "Rolls a dice with X sides, with an option to add on Y at the end of the roll",
               "!randomOWHero" : "Rolls a random Overwatch hero",
               "!score [string SPORT, string TEAM_NAME]": "Gets the current score of the game that TEAM_NAME is playing in",
               "X points to @[user]" : "Give points to a user, even though it's made up and the points don't matter",
               "!points" : "Get a list of all the points everyone has",
               "!betchallenge" : "Start a bet challenge",
               "!addbet [number of points you want to bet]" : "If a bet challenge has started, use this to enter the challenge and bet as many of your points as you want",
               "!roulette" : "Try your luck",
               "!cuteanimal" : "Scans selected subreddits for a cute animal photo"}


client = discord.Client()

#Bet Variables
canBetChallenge = True
canEnterBets = False
numEntries = 0
betEntries = []
betsEntered = {}

class Person:
    def __init__(self, name, id, excelCell):
        self.name = name
        self.id = id
        self.excelCell = excelCell

@client.event
async def on_message(message):
    # we do not want the bot to reply to itself
    if message.author == client.user:
        return
    else:
        if message.server.id == 'serverid':
            await trackMessages(message)
            await trackMentions(message)

    #Betting challenge functions
    global canBetChallenge
    global canEnterBets
    global numEntries
    global betEntries
    global betsEntered

    if(message.content.startswith('!addbet') and canEnterBets == True and message.author.id not in betEntries):
        entries = int(message.content.split(' ')[1])
        if entries < 1:
            await client.send_message(message.channel, PersonClassDict[message.author.id].name + "'s entry is not valid and has been thrown out")
            return

        wb = openpyxl.load_workbook(KnackchatSpreadsheetPath)
        ws = wb['Points']
        for cell in ws['A']:
            if cell.value == PersonClassDict[message.author.id].name:
                if ws['B' + str(cell.row)].value < entries:
                    entries = ws['B' + str(cell.row)].value
                    await client.send_message(message.channel, PersonClassDict[message.author.id].name + " put too many entries in, so I guess that means you're all in with " + str(ws['B' + str(cell.row)].value) + " entries")
            
                ws['B' + str(cell.row)].value -= entries

        for x in range(0, entries):
            betEntries.append(message.author.id)
        betsEntered[PersonClassDict[message.author.id].name] = entries

        numEntries += 1
        wb.save(KnackchatSpreadsheetPath)
        print('finished entering a user')


    if(message.content == '!betchallenge' and canBetChallenge and canEnterBets == False):
        canBetChallenge = False
        canEnterBets = True
        await timer(message.channel)
        canBetChallenge = True
        canEnterBets = False
        wb = openpyxl.load_workbook(KnackchatSpreadsheetPath)
        ws = wb['Points']
        
        if numEntries <= 1:
            for cell in ws['A']:
                if cell.value == PersonClassDict[message.author.id].name:
                    ws['B' + str(cell.row)].value += betsEntered[PersonClassDict[message.author.id].name]
                    await client.send_message(message.channel, "Not enough entries, your points have been returned")
        else:
            rand = random.randint(0, len(betEntries) - 1)
            for cell in ws['A']:
                if cell.value == PersonClassDict[betEntries[rand]].name:
                    ws['B' + str(cell.row)].value += len(betEntries)

            await client.send_message(message.channel, PersonClassDict[betEntries[rand]].name + " is the winner!")
            print('winner choosen')
        wb.save(KnackchatSpreadsheetPath)
        numEntries = 0
        betEntries.clear()
        betsEntered.clear()



    #Prints a list of commands and what they do
    if(message.content.startswith('!help') or message.content.startswith('!commands')):
        await getCommands(message.channel)

    #Get a cute animal photo
    if message.content == '!cuteanimal':
        await cuteAnimal(message.channel)


    #Discord Roulette
    if message.content == '!roulette':
        rand = random.randint(0, 5)
        if rand == 0:
            await client.send_message(message.channel, ':gun: BANG :gun:')
        else:
            await client.send_message(message.channel, ':hearts:')

    #Assign Points to User
    isAnInteger = False
    words = message.content.split(' ')
    try:
        int(words[0])
        isAnInteger = True
    except ValueError:
        isAnInteger = False

    if isAnInteger and ("points for" in message.content or "points to" in message.content or "point to" in message.content or "point for" in message.content) and len(message.raw_mentions) > 0:
        if message.raw_mentions[0] == message.author.id:
            wb = openpyxl.load_workbook(KnackchatSpreadsheetPath)
            ws = wb['Points']
            targetUser = PersonClassDict[message.author.id].name
            for cell in ws['A']:
                if cell.value == targetUser:
                    ws['B' + str(cell.row)].value -= 10

            wb.save(KnackchatSpreadsheetPath)
            print("Finished with Excel")
            await client.send_message(message.channel, "Nice try, -10 points for " + targetUser)
            return
        else:
            points = int(words[0])
            if points < 0:
                await client.send_message(message.channel, "ಠ_ಠ no.")
                return

            if points > 10:
                points = 10
            targetUser = PersonClassDict[message.raw_mentions[0]].name
            print(targetUser)
            authorUser = PersonClassDict[message.author.id].name
            print(authorUser)

            wb = openpyxl.load_workbook(KnackchatSpreadsheetPath)
            ws = wb['Points']
            for cell in ws['A']:
                if cell.value == targetUser:
                    print('added points')
                    ws['B' + str(cell.row)].value += points

            wb.save(KnackchatSpreadsheetPath)
            print("Finished with Excel")

            if points > 1:
                await client.send_message(message.channel, authorUser + " has given " + str(points) + " points to " + targetUser + "!")
            else:
                await client.send_message(message.channel, authorUser + " has given " + str(points) + " point to " + targetUser + "!")

    #Get a list of all the points users have
    if message.content == '!points':
        wb = openpyxl.load_workbook(KnackchatSpreadsheetPath)
        ws = wb['Points']
        result = ''
        firstTime = True
        for cell in ws['A']:
            if firstTime:
                firstTime = False
                continue
            else:
                result += "**" + cell.value + "**    " + str(ws['B' + str(cell.row)].value) + "\n"

        wb.save(KnackchatSpreadsheetPath)
        print("Finished with Excel")
        await client.send_message(message.channel, result)


    #Roll a dice of any size
    if message.content.startswith('!roll'):
        rollList = message.content.split(" ")
        if len(rollList) < 1:
            await client.send_message(message.channel, "Please enter a number of sides")
        else:
            sides = int(rollList[1])
            result = random.randint(1, sides)
            if len(rollList) == 3:
                result += int(rollList[2])
            await client.send_message(message.channel, "Roll Result: " + str(result))
      

    #Roll a random Overwatch Hero
    if message.content == '!randomOWHero':
        result = random.randint(0, len(owHeros) - 1)
        await client.send_message(message.channel, "Time To Play " + str(owHeros[result]) + "!")


    #Search for sports scores
    #Add ability for user to see the current match time
    if message.content.startswith('!score'):
        await getSportsScore(message)


@client.event
async def on_reaction_add(reaction: discord.Reaction, user):
    if reaction.message.server.id == 'serverid':
        print("new emoji added")
        wb = openpyxl.load_workbook(KnackchatSpreadsheetPath)
        ws = wb['Emoji Tracker']
        emojiName = ''
        if reaction.custom_emoji == False:
            emojiText = str(emoji.UNICODE_EMOJI_ALIAS[reaction.emoji]).split(':')
            emojiName = emojiText[1]
        else:
            emojiName = reaction.emoji.name

        for cell in ws['A']:
            if cell.value == emojiName:
                ws['B' + str(cell.row)].value += 1
                break
            if cell.value == None:
                newCell = ws['A' + str(cell.row)]
                newCell.value = emojiName
                ws['B' + str(cell.row)].value = 1
                break
        else:
            newCell = ws['A' + str(cell.row + 1)]
            newCell.value = emojiName
            ws['B' + str(cell.row + 1)].value = 1

        wb.save(KnackchatSpreadsheetPath)

@client.event
async def on_ready():
    rand = random.randint(0, len(gameList) - 1);
    await client.change_presence(game=discord.Game(name=gameList[rand], type=0))
    #await client.change_presence(game=discord.Game(name="Baby Come Knack", type=2));
    await createPersonClasses()

    print('Logged in as')
    print(client.user.name)
    print(client.user.id)
    print('------')

async def createPersonClasses():
    # populate the PersonClassDict here
    # person = person = Person('personname', 'persondiscordip', 'excelNum')
    # PersonClassDict[person.id] = person

    print('All person clsses created')

# Get bot commands
async def getCommands(channel):
    print(channel)
    commands = ''
    for x, y in commandDict.items():
        commands += ("**" + x + "** - " + y + "\n")

    await client.send_message(channel, commands)

# Get the score of a sports game based off a team name
async def getSportsScore(mess):
    subStrings = mess.content.split(" ", 2);
    sport = subStrings[1].lower()
    teamName = subStrings[2];
    await client.send_message(mess.channel, "Searching for the " + teamName + " game...")

    matches = sports.all_matches()
    specificSport = matches[sport]
    otherTeam = ""
    for x in range(len(specificSport)):
        print("home: " + str(specificSport[x].home_team) + " | away: " + str(specificSport[x].away_team))
        if(teamName == specificSport[x].home_team or teamName == specificSport[x].away_team):
            print('Found team')
            await client.send_message(mess.channel, str(specificSport[x]) + " | " + str(specificSport[x].match_time))
            return

    await client.send_message(mess.channel, "The " + teamName + " are currently not playing, please try again later.")

# for betchallenge
async def timer(channel):
    await client.send_message(channel, "You have 30 seconds to place your bets kids!")
    await asyncio.sleep(20)
    await client.send_message(channel, "10 seconds left...")
    await asyncio.sleep(20)

    await client.send_message(channel, "Entries closed.")

# Search for a random photo of a cute animal
async def cuteAnimal(channel):
    print('Searching for a cute animal photo...')
    loopSubreddits = True
    while loopSubreddits:
        print('Searching new post...')
        rand = random.randint(0, len(cuteSubreddits) - 1)
        subreddit = reddit.subreddit(cuteSubreddits[rand])
        posts = [post for post in subreddit.hot(limit=50)]
        rand = random.randint(0, 49)
        random_post = posts[rand]
        if not random_post.stickied and not random_post.over_18:
            checkIfValid = random_post.url.split('.')
            if len(checkIfValid) > 2:
                lastIndex = len(checkIfValid) - 1
                if checkIfValid[lastIndex] == 'jpg' or checkIfValid[lastIndex] == 'gif' or checkIfValid[lastIndex] == 'gifv' or checkIfValid[lastIndex] == 'png':
                    loopSubreddits = False
                    print('Valid post found on subreddit ' + str(subreddit))
                    await client.send_message(channel, random_post.title + "\n" + random_post.url)
                    return
    await None

# Track number of user messages
async def trackMessages(mess):
    if mess.author.id not in PersonClassDict:
        print('====== Message author was not found in dictionary: ' + mess.author.id + ' ======')
        return

    wb = openpyxl.load_workbook(KnackchatSpreadsheetPath)

    #Average Words Per Message
    ws = wb['AverageWordsPerMessage']

    targetCellNum = PersonClassDict[mess.author.id].excelCell
    
    ws['C' + targetCellNum].value += 1

    currentAvg = ws['B' + targetCellNum].value
    totalMessages = ws['C' + targetCellNum].value

    sentenceSplit = mess.content.split(' ')
    messageLength = len(sentenceSplit)

    newAvg = ((currentAvg * totalMessages) + messageLength) / (totalMessages)
    newAvg = round(newAvg, 1)
    ws['B' + targetCellNum].value = newAvg

    print('new message tracked')
    wb.save(KnackchatSpreadsheetPath)

# Track number of user mentions
async def trackMentions(message):
    if message.author.id not in PersonClassDict:
        print('====== Message author was not found in dictionary: ' + mess.author.id + ' ======')
        return

    #check if any user has been mentioned
    usersMentioned = message.raw_mentions
    if(len(usersMentioned) < 1):
        return

    #open up the spreadsheet
    wb = openpyxl.load_workbook(KnackchatSpreadsheetPath)
    ws = wb['UserMentions']

    #Find the author
    targetCellNum = PersonClassDict[message.author.id].excelCell
    #Set new total of users mentions
    ws['C' + targetCellNum].value += len(usersMentioned)

    #Add one mention to each user mentioned
    for x in range(0, len(usersMentioned)):
        person = PersonClassDict[usersMentioned[x]]
        ws['B' + person.excelCell].value += 1

        #Add specific user mentioned to the author's dictionary
        currentDict = ast.literal_eval(ws['D' + targetCellNum].value)
        if person.name in currentDict:
            currentDict[person.name] += 1
        else:
            currentDict[person.name] = 1
        ws['D' + targetCellNum].value = str(currentDict)

    print('User(s) has been mentioned and tracked')

    wb.save(KnackchatSpreadsheetPath)

#WIP - find and replace any of this function in every other function
def searchWSForCell(wb, wsName, searchCol, cellTargetValue):
    ws = wb[wsName]
    for cell in ws[searchCol]:
        if cell.value == cellTargetValue:
            return cell

    return None

client.run(TOKEN)